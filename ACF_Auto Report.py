import os
import re
import shutil
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import sys

# ============================================================
# 辅助函数
# ============================================================

def parse_folder_name(folder_name):
    parts = folder_name.strip().split()
    if len(parts) < 5:
        print(f"⚠️ 警告：文件夹名称格式不正确，需要至少5个部分")
        print(f"   当前：{folder_name}")
        return None
    return {
        "专案": parts[0],
        "阶段": parts[1],
        "Config": parts[2],
        "机台号": parts[3],
        "报告类型": parts[4]
    }

def get_first_file_in_folder(folder_path, extensions=None):
    if not os.path.exists(folder_path):
        return None
    files = [
        f for f in os.listdir(folder_path)
        if os.path.isfile(os.path.join(folder_path, f))
        and not f.startswith('~$')
        and not f.startswith('.')
    ]
    if extensions:
        files = [f for f in files if any(f.lower().endswith(ext) for ext in extensions)]
    if files:
        return os.path.join(folder_path, files[0])
    return None

def extract_revision_from_file(file_path, search_pattern):
    if not file_path or not os.path.exists(file_path):
        return None
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        match = re.search(search_pattern, content)
        if match:
            return match.group(1)
        return None
    except Exception as e:
        print(f"   ⚠️ 读取文件失败：{e}")
        return None

def safe_set_cell_value(ws, row, col, value):
    for merged_range in ws.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and col in range(merged_range.min_col, merged_range.max_col + 1):
            ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
            return
    ws.cell(row=row, column=col).value = value

def get_merged_cell_range(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if row in range(merged_range.min_row, merged_range.max_row + 1) and col in range(merged_range.min_col, merged_range.max_col + 1):
            return (merged_range.min_row, merged_range.max_row, merged_range.min_col, merged_range.max_col)
    return None

# ============================================================
# 主函数
# ============================================================
def update_module_integrity_report(wb, data_folder_path, 报告类型):
    """
    更新 Module integrity Sheet：
    - MBO：清空 D-M 列 12/14/16/18/20 行的图片，从源图按行查找并插入（锚定到 D 列开始）
    - PBO：清空 D-M 列 12/14/16/18/20 行的图片，并将这些单元格填入 "NA"
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO
    from PIL import Image

    print("\n" + "=" * 60)
    print("📝 更新 Module integrity Sheet")
    print("=" * 60)

    # 1. 定位目标 Sheet
    sheet_names = wb.sheetnames
    target_sheet_name = None
    for name in sheet_names:
        if "module integrity" in name.lower():
            target_sheet_name = name
            break
    if target_sheet_name is None:
        print("   ⚠️ 未找到 'Module integrity' Sheet，跳过更新")
        return

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 2. 目标行和列范围
    target_rows = [12, 14, 16, 18, 20]
    col_start = 4  # D
    col_end = 13   # M

    # 3. 清空目标区域的图片（所有目标行，D-M 列）
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if img_row in target_rows and col_start <= img_col <= col_end:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 4. PBO 模式：清空并填入 "NA"，直接返回
    if 报告类型 == "PBO":
        print("   📝 PBO 模式：填入 'NA'")
        for row in target_rows:
            for col in range(col_start, col_end + 1):
                safe_set_cell_value(ws_target, row, col, "NA")
        print("   ✅ PBO 更新完成")
        return

    # 5. MBO 模式：从 IPQC Data 复制图片
    print("   📝 MBO 模式：从源图复制图片")

    # 定位 IPQC Data 文件
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过图片复制")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        # 模糊匹配源 Sheet 名称（包含 "module integrity"）
        src_sheet_name = None
        for name in wb_ipqc.sheetnames:
            if "module integrity" in name.lower():
                src_sheet_name = name
                break
        if src_sheet_name is None:
            print(f"   ⚠️ IPQC Data 文件中没有包含 'Module integrity' 的 Sheet，跳过图片复制")
            print(f"   📋 可用的 Sheet：{wb_ipqc.sheetnames}")
            wb_ipqc.close()
            return
        ws_source = wb_ipqc[src_sheet_name]
        print(f"   ✅ 使用源 Sheet：'{src_sheet_name}'")
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 6. 收集源图片：放宽行范围（-1,+1），不限制列范围，只要在数据行附近
    source_images = []  # (src_row, src_col, img_data)
    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                # 检查是否属于某个目标行（±1 范围）
                for target_row in target_rows:
                    if target_row - 1 <= img_row <= target_row + 1:
                        try:
                            img_data = img._data()
                            source_images.append((img_row, img_col, img_data))
                        except Exception as e:
                            print(f"      ⚠️ 无法读取源图片数据（行{img_row}列{get_column_letter(img_col)}）：{e}")
                        break  # 一旦匹配到就跳出，避免重复

    if not source_images:
        print("   ⚠️ 没有找到需要复制的图片")
        wb_ipqc.close()
        return

    # 按行、列排序
    source_images.sort(key=lambda x: (x[0], x[1]))
    print(f"   📷 找到 {len(source_images)} 张源图片")

    # 7. 将图片分配到对应的目标行（每个目标行可能有多张图片）
    images_by_row = {}
    for src_row, src_col, img_data in source_images:
        # 确定属于哪个目标行（取最近的）
        target_row = None
        for tr in target_rows:
            if tr - 1 <= src_row <= tr + 1:
                target_row = tr
                break
        if target_row is None:
            continue
        if target_row not in images_by_row:
            images_by_row[target_row] = []
        images_by_row[target_row].append((src_col, img_data))

    # 8. 获取列宽（D 列参考）
    col_width = ws_target.column_dimensions[get_column_letter(col_start)].width
    if col_width is None:
        col_width = 12
    col_width_px = col_width * 8

    # 9. 插入图片到对应行（从 D 列开始依次放置）
    print("   📝 插入图片...")
    for target_row, imgs in images_by_row.items():
        # 按列排序
        imgs.sort(key=lambda x: x[0])
        # 最多放到 M 列
        max_cols = col_end - col_start + 1
        for idx, (src_col, img_data) in enumerate(imgs[:max_cols]):
            target_col = col_start + idx
            try:
                pil_img = Image.open(BytesIO(img_data))
                # 移除 EXIF
                pil_img = pil_img.convert('RGB')
                output = BytesIO()
                pil_img.save(output, format='PNG')
                output.seek(0)
                new_img = XLImage(output)

                # 获取行高
                row_height_pts = ws_target.row_dimensions[target_row].height
                if row_height_pts is None:
                    row_height_pts = 15
                row_height_px = row_height_pts * 1.3333

                # 尺寸：宽度固定，高度填满
                new_img.width = col_width_px * 0.95
                new_img.height = row_height_px * 0.95

                ws_target.add_image(new_img, f"{get_column_letter(target_col)}{target_row}")
                print(f"      ✅ 插入图片到 {get_column_letter(target_col)}{target_row}（来自行{src_row}列{get_column_letter(src_col)}）")
            except Exception as e:
                print(f"      ⚠️ 插入图片到 {get_column_letter(target_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Module integrity 更新完成")
    print("=" * 60)

def update_fai_report(wb, data_folder_path):
    """
    从 IPQC Data 的 FAI Sheet 复制 D10:G19 数据到报告的 FAI Sheet。
    """
    print("\n" + "=" * 60)
    print("📝 更新 FAI Sheet")
    print("=" * 60)

    # 1. 定位目标 Sheet（容错）
    sheet_names = wb.sheetnames
    target_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() == "fai":
            target_sheet_name = name
            break
    if target_sheet_name is None:
        print("   ⚠️ 未找到 'FAI' Sheet，跳过更新")
        return

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 2. 定位 IPQC Data 中的对应 Sheet
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        src_sheet_name = None
        for name in wb_ipqc.sheetnames:
            if name.strip().lower() == "fai":
                src_sheet_name = name
                break
        if src_sheet_name is None:
            print(f"   ⚠️ IPQC Data 文件中没有 'FAI' Sheet，跳过更新")
            wb_ipqc.close()
            return
        ws_source = wb_ipqc[src_sheet_name]
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 3. 复制数据（D10:G19）
    print("   📝 复制 D10:G19...")
    for row in range(10, 20):
        for col in range(4, 8):  # D=4, G=7
            val = ws_source.cell(row=row, column=col).value
            ws_target.cell(row=row, column=col, value=val)
    print("   ✅ 复制完成")

    wb_ipqc.close()
    print("   ✅ FAI 更新完成")
    print("=" * 60)

def update_break_mode_report(wb, data_folder_path):
    """
    更新 Break Mode Sheet：
    1. 清空 C8:BK21 和 C22:CC69 的数据和图片
    2. 从 IPQC Data 复制文本数据到相同区域
    3. 从源区域按顺序提取图片，插入到锚定点列表
    4. 插入图片前，设置 D:CC 列宽为 2.56
    5. 图片尺寸：D13, AH13 宽度=合并单元格宽度*30%，高度=合并单元格高度
                    其他图片 宽度=合并单元格宽度*50%，高度=合并单元格高度
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO
    from PIL import Image

    print("\n" + "=" * 60)
    print("📝 更新 Break Mode Sheet")
    print("=" * 60)

    # 1. 定位目标 Sheet（容错）
    sheet_names = wb.sheetnames
    target_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() == "break mode":
            target_sheet_name = name
            break
    if target_sheet_name is None:
        print("   ⚠️ 未找到 'Break Mode' Sheet，跳过更新")
        return

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 2. 定位 IPQC Data 中的对应 Sheet
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        src_sheet_name = None
        for name in wb_ipqc.sheetnames:
            if name.strip().lower() == "break mode":
                src_sheet_name = name
                break
        if src_sheet_name is None:
            print(f"   ⚠️ IPQC Data 文件中没有 'Break Mode' Sheet，跳过更新")
            wb_ipqc.close()
            return
        ws_source = wb_ipqc[src_sheet_name]
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 3. 定义区域
    regions = [
        (8, 21, 3, 63),   # C8:BK21 (C=3, BK=63)
        (22, 69, 3, 81)   # C22:CC69 (C=3, CC=81)
    ]

    # 4. 清空目标区域的数据和图片
    print("   📝 清空数据区域...")
    for start_row, end_row, start_col, end_col in regions:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                safe_set_cell_value(ws_target, row, col, None)

    # 清空旧图片（整个区域）
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                # 检查是否在区域内
                in_region = False
                for r_start, r_end, c_start, c_end in regions:
                    if r_start <= img_row <= r_end and c_start <= img_col <= c_end:
                        in_region = True
                        break
                if in_region:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 5. 复制文本数据
    print("   📝 复制文本数据...")
    for start_row, end_row, start_col, end_col in regions:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                val = ws_source.cell(row=row, column=col).value
                safe_set_cell_value(ws_target, row, col, val)
        print(f"      行 {start_row}-{end_row} 列 {get_column_letter(start_col)}-{get_column_letter(end_col)} 复制完成")

    # 6. 收集源区域内的所有图片（按顺序：行优先，列其次）
    source_images = []  # 每个元素为 (img_data, src_row, src_col)
    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                # 检查是否在区域内
                in_region = False
                for r_start, r_end, c_start, c_end in regions:
                    if r_start <= img_row <= r_end and c_start <= img_col <= c_end:
                        in_region = True
                        break
                if in_region:
                    try:
                        img_data = img._data()
                        source_images.append((img_data, img_row, img_col))
                    except Exception as e:
                        print(f"      ⚠️ 无法读取源图片数据（行{img_row}列{get_column_letter(img_col)}）：{e}")
                        continue

    # 按行、列排序
    source_images.sort(key=lambda x: (x[1], x[2]))
    print(f"   📷 找到 {len(source_images)} 张源图片")

    # 7. 目标锚定点列表（按插入顺序）
    anchor_cells = [
        (13, 4),   # D13
        (13, 34),  # AH13
        (25, 4),   # D25
        (25, 43),  # AQ25
        (37, 4),   # D37
        (37, 43),  # AQ37
        (49, 4),   # D49
        (49, 43),  # AQ49
        (61, 4),   # D61
        (61, 43),  # AQ61
    ]

    # 8. 在插入图片前，设置 D:CC 列宽为 2.56
    print("   📏 设置 D:CC 列宽为 2.56...")
    for col in range(4, 82):  # D=4, CC=81
        col_letter = get_column_letter(col)
        ws_target.column_dimensions[col_letter].width = 2.56
    print("      ✅ 列宽设置完成")

    # 定义辅助函数：获取合并单元格的总宽度（像素）
    def get_merged_width(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and col in range(merged_range.min_col, merged_range.max_col + 1):
                total_width = 0
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    w = ws.column_dimensions[get_column_letter(c)].width
                    if w is None:
                        w = 2.56  # 默认列宽
                    total_width += w
                return total_width * 8  # 转换为像素（1字符≈8像素）
        # 非合并单元格，取单列宽
        w = ws.column_dimensions[get_column_letter(col)].width
        if w is None:
            w = 2.56
        return w * 8

    def get_merged_height(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and col in range(merged_range.min_col, merged_range.max_col + 1):
                total_height = 0
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    h = ws.row_dimensions[r].height
                    if h is None:
                        h = 15
                    total_height += h
                return total_height * 1.3333
        # 非合并单元格，取单行高
        h = ws.row_dimensions[row].height
        if h is None:
            h = 15
        return h * 1.3333

    # 插入图片
    print("   📝 插入图片（锚定到指定位置）...")
    for idx, (target_row, target_col) in enumerate(anchor_cells):
        if idx >= len(source_images):
            print(f"      ⚠️ 图片数量不足，只插入 {len(source_images)} 张")
            break
        img_data, src_row, src_col = source_images[idx]
        try:
            pil_img = Image.open(BytesIO(img_data))
            # 移除 EXIF 信息，避免旋转
            pil_img = pil_img.convert('RGB')
            output = BytesIO()
            pil_img.save(output, format='PNG')
            output.seek(0)
            new_img = XLImage(output)

            # 获取合并单元格尺寸
            merged_width_px = get_merged_width(ws_target, target_row, target_col)
            merged_height_px = get_merged_height(ws_target, target_row, target_col)

            # 根据锚定位置决定缩放比例
            if (target_row == 13 and target_col == 4) or (target_row == 13 and target_col == 34):
                scale = 0.30  # D13, AH13 宽度30%
            else:
                scale = 0.50  # 其他图片宽度50%

            # 设置图片尺寸：宽度按比例，高度填满合并单元格
            img_width = merged_width_px * scale
            img_height = merged_height_px  # 高度100%

            new_img.width = img_width
            new_img.height = img_height

            ws_target.add_image(new_img, f"{get_column_letter(target_col)}{target_row}")
            print(f"      ✅ 插入图片到 {get_column_letter(target_col)}{target_row}（缩放比例 {scale*100:.0f}%）")
        except Exception as e:
            print(f"      ⚠️ 插入图片到 {get_column_letter(target_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Break Mode 更新完成")
    print("=" * 60)

def update_acf_squeeze_out_report(wb, data_folder_path):
    """
    更新 ACF squeeze out Sheet：
    - 数据列映射：源 O,P,Q,S,T,U,W,X,Y → 报告 Q,R,S,U,V,W,Y,Z,AA（行 7-16）
    - 图片按列处理：P, T, X 列（源 N, R, V → 目标 P, T, X）
    - 每列独立行范围匹配（±1行），防止跨列干扰
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    print("\n" + "=" * 60)
    print("📝 更新 ACF squeeze out Sheet（按列处理图片）")
    print("=" * 60)

    # 1. 定位 Sheet
    sheet_names = wb.sheetnames
    target_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() == "acf squeeze out":
            target_sheet_name = name
            break
    if target_sheet_name is None:
        print("   ⚠️ 未找到 'ACF squeeze out' Sheet，跳过更新")
        return

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 2. 定位 IPQC Data 中的对应 Sheet
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        src_sheet_name = None
        for name in wb_ipqc.sheetnames:
            if name.strip().lower() == "acf squeeze out":
                src_sheet_name = name
                break
        if src_sheet_name is None:
            print(f"   ⚠️ IPQC Data 文件中没有 'ACF squeeze out' Sheet，跳过更新")
            wb_ipqc.close()
            return
        ws_source = wb_ipqc[src_sheet_name]
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 3. 列配置
    # 数据映射：源 O,P,Q,S,T,U,W,X,Y → 报告 Q,R,S,U,V,W,Y,Z,AA
    source_data_cols = [15, 16, 17, 19, 20, 21, 23, 24, 25]
    target_data_cols = [17, 18, 19, 21, 22, 23, 25, 26, 27]
    # 图片列映射：源 N,R,V → 报告 P,T,X（对应关系）
    image_mapping = [
        (14, 16),  # N → P
        (18, 20),  # R → T
        (22, 24)   # V → X
    ]
    target_rows = list(range(7, 17))

    # 4. 清空旧图片（目标图片列，行 7-16）
    target_image_cols = [16, 20, 24]
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if img_row in target_rows and img_col in target_image_cols:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 5. 复制数据
    print("   📝 复制数据（按列映射）...")
    for row in target_rows:
        for src_col, tgt_col in zip(source_data_cols, target_data_cols):
            val = ws_source.cell(row=row, column=src_col).value
            ws_target.cell(row=row, column=tgt_col, value=val)
        print(f"      行 {row} 更新完成")

    # 6. 按列处理图片
    # 准备列宽（参考 P 列）
    ref_col = target_image_cols[0]
    ref_col_letter = get_column_letter(ref_col)
    col_width = ws_target.column_dimensions[ref_col_letter].width
    if col_width is None:
        col_width = 12
    col_width_px = col_width * 7.5

    print("   📝 按列处理图片...")
    for src_col, tgt_col in image_mapping:
        print(f"     处理列：源 {get_column_letter(src_col)} → 目标 {get_column_letter(tgt_col)}")
        # 收集该源列的所有图片（放宽行范围 5-20）
        col_images = []  # (源行, 图片数据)
        if hasattr(ws_source, '_images') and ws_source._images:
            for img in ws_source._images:
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    from_cell = img.anchor._from
                    img_row = from_cell.row + 1
                    img_col = from_cell.col + 1
                    if 5 <= img_row <= 20 and img_col == src_col:
                        try:
                            img_data = img._data()
                            col_images.append((img_row, img_data))
                        except Exception as e:
                            print(f"         ⚠️ 无法读取源图片数据（行{img_row}）：{e}")
                            continue

        if not col_images:
            print(f"         ⚠️ 列 {get_column_letter(src_col)} 没有找到图片")
            continue

        # 按源行排序
        col_images.sort(key=lambda x: x[0])

        # 匹配到目标行：每个目标行在 ±1 范围内查找未使用的图片
        used = [False] * len(col_images)
        for target_row in target_rows:
            # 查找候选
            candidates = []
            for idx, (src_row, img_data) in enumerate(col_images):
                if not used[idx] and (target_row - 1 <= src_row <= target_row + 1):
                    candidates.append((idx, img_data))
            if candidates:
                # 取第一个（通常只有一个）
                idx, img_data = candidates[0]
                used[idx] = True
                # 插入图片到目标单元格
                try:
                    img_stream = BytesIO(img_data)
                    new_img = XLImage(img_stream)

                    row_height_pts = ws_target.row_dimensions[target_row].height
                    if row_height_pts is None:
                        row_height_pts = 15
                    row_height_px = row_height_pts * 1.3333

                    new_img.width = col_width_px * 0.95
                    new_img.height = row_height_px * 0.95

                    ws_target.add_image(new_img, f"{get_column_letter(tgt_col)}{target_row}")
                    print(f"         ✅ 插入图片到 {get_column_letter(tgt_col)}{target_row}")
                except Exception as e:
                    print(f"         ⚠️ 插入图片到 {get_column_letter(tgt_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ ACF squeeze out 更新完成（按列处理图片）")
    print("=" * 60)

def update_lamination_placement_images(wb, data_folder_path):
    """
    从 IPQC Data 的 Lamination placement Sheet 复制 Q 列图片到报告，
    采用顺序映射：源图片按行排序，目标行按行排序，一一对应。
    清空旧图片后，复制新图片。
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    print("\n" + "=" * 60)
    print("📝 更新 Lamination placement Q 列图片（顺序映射）")
    print("=" * 60)

    # 1. 定位目标 Sheet（兼容名称带空格）
    sheet_names = wb.sheetnames
    target_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() == "lamination placement":
            target_sheet_name = name
            break
    if target_sheet_name is None:
        print("   ⚠️ 未找到 'Lamination placement' Sheet，跳过图片更新")
        return

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 2. 定位 IPQC Data 中的对应 Sheet
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过图片更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        src_sheet_name = None
        for name in wb_ipqc.sheetnames:
            if name.strip().lower() == "lamination placement":
                src_sheet_name = name
                break
        if src_sheet_name is None:
            print(f"   ⚠️ IPQC Data 文件中没有 'Lamination placement' Sheet，跳过图片更新")
            wb_ipqc.close()
            return
        ws_source = wb_ipqc[src_sheet_name]
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    q_col = 17  # Q列
    q_col_letter = get_column_letter(q_col)

    # 3. 清空目标 Sheet 中 Q 列的旧图片
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_col = from_cell.col + 1
                if img_col == q_col:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张 Q 列旧图片")

    # 4. 收集源 Q 列图片（按行排序）
    source_images = []
    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if img_col == q_col:
                    try:
                        img_data = img._data()
                        source_images.append((img_row, img_data))
                    except Exception as e:
                        print(f"      ⚠️ 无法读取源图片数据（行{img_row}）：{e}")
                        continue

    if not source_images:
        print("   ⚠️ 没有找到需要复制的 Q 列图片")
        wb_ipqc.close()
        return

    # 按行排序
    source_images.sort(key=lambda x: x[0])
    print(f"   📷 找到 {len(source_images)} 张 Q 列源图片，行号范围：{source_images[0][0]} ~ {source_images[-1][0]}")

    # 5. 确定目标行（报告中 Q 列需要放置图片的行）
    # 从数据区域推断：E:J 列，行 8-17 和 30-39，所以 Q 列图片也对应这些行
    target_rows = list(range(8, 18)) + list(range(30, 40))
    # 如果源图片数量超过目标行数，只取前 target_rows 个；如果不足，用所有源图片
    max_images = min(len(source_images), len(target_rows))
    source_images = source_images[:max_images]
    target_rows = target_rows[:max_images]

    print(f"   📌 目标行（前 {len(target_rows)} 个）：{target_rows[:5]}...{target_rows[-5:] if len(target_rows)>5 else ''}")

    # 6. 获取 Q 列宽度（用于调整图片大小）
    q_col_width = ws_target.column_dimensions[q_col_letter].width
    if q_col_width is None:
        q_col_width = 12
    col_width_px = q_col_width * 8

    # 7. 一一映射：第 i 个源图片 → 第 i 个目标行
    for i, (src_row, img_data) in enumerate(source_images):
        target_row = target_rows[i]
        try:
            img_stream = BytesIO(img_data)
            new_img = XLImage(img_stream)

            # 获取目标行高
            row_height_pts = ws_target.row_dimensions[target_row].height
            if row_height_pts is None:
                row_height_pts = 15
            row_height_px = row_height_pts * 1.3333

            new_img.width = col_width_px * 0.99
            new_img.height = row_height_px * 0.99

            ws_target.add_image(new_img, f"{q_col_letter}{target_row}")
            print(f"      ✅ 插入图片到 {q_col_letter}{target_row}（来自源行 {src_row}）")
        except Exception as e:
            print(f"      ⚠️ 插入图片到 {q_col_letter}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Lamination placement Q 列图片更新完成（顺序映射）")
    print("=" * 60)

def update_lamination_placement_report(wb, data_folder_path):
    """
    更新 Lamination placement Sheet
    数据范围：E:J 列，行 8-17 和 30-39
    """
    print("\n" + "=" * 60)
    print("📝 更新 Lamination placement 数据")
    print("=" * 60)

    # 1. 确定目标 Sheet 名称（去除空格匹配）
    sheet_names = wb.sheetnames
    target_sheet_name = None

    # 先尝试精确匹配（带空格）
    if "Lamination placement" in sheet_names:
        target_sheet_name = "Lamination placement"
    elif "Lamination placement " in sheet_names:
        target_sheet_name = "Lamination placement "
    else:
        # 如果找不到，尝试去除空格后匹配
        for name in sheet_names:
            if name.strip() == "Lamination placement":
                target_sheet_name = name
                break

    if target_sheet_name is None:
        # 如果仍然找不到，尝试找 Compressed ball diameter 后的 Sheet
        if "Compressed ball diameter" in sheet_names:
            idx = sheet_names.index("Compressed ball diameter")
            if idx + 1 < len(sheet_names):
                target_sheet_name = sheet_names[idx + 1]
                print(f"   ✅ 使用 'Compressed ball diameter' 后的 Sheet：'{target_sheet_name}'")
            else:
                print("   ⚠️ 'Compressed ball diameter' 已是最后一个 Sheet")
                return
        else:
            print("   ⚠️ 未找到 'Lamination placement' 或 'Compressed ball diameter'")
            return

    print(f"   ✅ 目标 Sheet：'{target_sheet_name}'")

    try:
        ws_target = wb[target_sheet_name]
    except KeyError:
        print(f"   ⚠️ 报告中找不到 Sheet '{target_sheet_name}'")
        return

    # 其余逻辑不变...

    # 2. 定位 IPQC Data 中的对应 Sheet
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if target_sheet_name in wb_ipqc.sheetnames:
            ws_source = wb_ipqc[target_sheet_name]
        else:
            print(f"   ⚠️ IPQC Data 文件中没有 Sheet '{target_sheet_name}'，跳过更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 3. 复制数据（E:J 列，行 8-17 和 30-39）
    ranges = [
        (8, 17),
        (30, 39)
    ]
    col_start = 5  # E
    col_end = 10   # J

    print(f"   📝 复制数据（E-J 列，行 8-17 和 30-39）到 Sheet '{target_sheet_name}'...")
    for start_row, end_row in ranges:
        for row in range(start_row, end_row + 1):
            for col in range(col_start, col_end + 1):
                val = ws_source.cell(row=row, column=col).value
                ws_target.cell(row=row, column=col, value=val)  # 完全覆盖，包括 None
        print(f"      行 {start_row}-{end_row} 更新完成")

    wb_ipqc.close()
    print(f"   ✅ Lamination placement（{target_sheet_name}）数据更新完成")
    print("=" * 60)

def update_compressed_ball_diameter_block2(wb, data_folder_path):
    """
    处理 Compressed ball diameter Sheet 的第二个区块：
    - 数据行：162, 181, 200, 219（D:AA 列，每4列合并为一个数据块）
    - 图片行：144, 153, 163, 172, 182, 191, 201, 210（D:AA 列，每4列一个图片）
    支持合并单元格的尺寸计算。
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    print("\n" + "=" * 60)
    print("📝 更新 Compressed ball diameter 第二个区块")
    print("=" * 60)

    try:
        ws_target = wb["Compressed ball diameter"]
    except KeyError:
        print("   ⚠️ 未找到 'Compressed ball diameter' Sheet，跳过区块2更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过区块2更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Compressed ball diameter" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Compressed ball diameter"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Compressed ball diameter' Sheet，跳过更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # ============================================================
    # 1. 数据复制（D:AA 列，行 162,181,200,219）
    # ============================================================
    data_rows = [162, 181, 200, 219]
    col_start = 4  # D
    col_end = 27   # AA

    print("   📝 复制数据（D-AA 列，行 162,181,200,219）...")
    for row in data_rows:
        for col in range(col_start, col_end + 1):
            val = ws_source.cell(row=row, column=col).value
            if val == 0 or val == 0.0:
                ws_target.cell(row=row, column=col, value=None)
            else:
                ws_target.cell(row=row, column=col, value=val)
        print(f"      行 {row} 更新完成")

    # ============================================================
    # 2. 定义合并单元格尺寸计算函数（复用）
    # ============================================================
    def get_merged_cell_total_height(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_height = 0
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    h = ws.row_dimensions[r].height
                    if h is None:
                        h = 15
                    total_height += h
                return total_height * 1.3333
        h = ws.row_dimensions[row].height
        if h is None:
            h = 15
        return h * 1.3333

    def get_merged_cell_total_width(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_width = 0
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    w = ws.column_dimensions[get_column_letter(c)].width
                    if w is None:
                        w = 12
                    total_width += w
                return total_width * 8
        w = ws.column_dimensions[get_column_letter(col)].width
        if w is None:
            w = 12
        return w * 8

    # ============================================================
    # 3. 图片复制
    # ============================================================
    target_image_rows = [144, 153, 163, 172, 182, 191, 201, 210]
    # 锚定列：从 D 开始，步长4（4列合并为一个单元格）
    anchor_cols = list(range(4, 28, 4))  # 4,8,12,16,20,24

    # 清除旧图片（目标行，D-AA 列）
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if img_row in target_image_rows and 4 <= img_col <= 27:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 收集源图片（从 IPQC Data，目标行 ±1 行，D-AA 列）
    source_images = {row: [] for row in target_image_rows}
    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                src_row = from_cell.row + 1
                src_col = from_cell.col + 1
                if 143 <= src_row <= 211 and 4 <= src_col <= 27:
                    for target_row in target_image_rows:
                        if target_row - 1 <= src_row <= target_row + 1:
                            source_images[target_row].append((src_col, img))
                            break

    # 插入图片
    print("   📝 插入图片（D-AA 列，每4列一组）...")
    for target_row, imgs in source_images.items():
        if not imgs:
            print(f"      ⚠️ 目标行 {target_row} 没有找到图片")
            continue
        imgs.sort(key=lambda x: x[0])  # 按列排序
        used_cols = anchor_cols[:len(imgs)]
        for idx, (src_col, img_obj) in enumerate(imgs):
            target_col = used_cols[idx]
            try:
                img_data = img_obj._data()
                img_stream = BytesIO(img_data)
                new_img = XLImage(img_stream)

                # 计算合并单元格的总宽度（4列）
                # 注意：目标单元格可能是合并列，也可能是4列合并，我们用 get_merged_cell_total_width 来处理
                col_width_px = get_merged_cell_total_width(ws_target, target_row, target_col)
                # 计算合并单元格的总高度（可能是多行合并）
                row_height_px = get_merged_cell_total_height(ws_target, target_row, target_col)

                new_img.width = col_width_px * 0.99
                new_img.height = row_height_px * 0.99
                ws_target.add_image(new_img, f"{get_column_letter(target_col)}{target_row}")
                print(f"      ✅ 插入图片到 {get_column_letter(target_col)}{target_row}")
            except Exception as e:
                print(f"      ⚠️ 插入图片到 {get_column_letter(target_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Compressed ball diameter 第二个区块更新完成")
    print("=" * 60)

def update_compressed_ball_diameter_images(wb, data_folder_path):
    """
    使用 openpyxl 从 IPQC Data 的 Compressed ball diameter Sheet 复制图片。
    目标行：102, 111, 120, 129
    查找范围：目标行 ±1 行，C-CG 列
    锚定列：D, G, J, ...（间隔3列），最大到 CD 列
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    print("\n" + "=" * 60)
    print("📝 更新 Compressed ball diameter 图片（改进版）")
    print("=" * 60)

    try:
        ws_target = wb["Compressed ball diameter"]
    except KeyError:
        print("   ⚠️ 未找到 'Compressed ball diameter' Sheet，跳过图片更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过图片更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Compressed ball diameter" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Compressed ball diameter"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Compressed ball diameter' Sheet，跳过图片更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 1. 清除旧图片（D-CF 列，行 100-140）
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if 100 <= img_row <= 140 and 4 <= img_col <= 84:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 2. 定义尺寸计算函数（支持合并单元格）
    def get_merged_cell_total_height(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_height = 0
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    h = ws.row_dimensions[r].height
                    if h is None:
                        h = 15
                    total_height += h
                return total_height * 1.3333
        h = ws.row_dimensions[row].height
        if h is None:
            h = 15
        return h * 1.3333

    def get_merged_cell_total_width(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_width = 0
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    w = ws.column_dimensions[get_column_letter(c)].width
                    if w is None:
                        w = 12
                    total_width += w
                return total_width * 8
        w = ws.column_dimensions[get_column_letter(col)].width
        if w is None:
            w = 12
        return w * 8

    # 3. 目标行与锚定列（D, G, J, ... 间隔3列）
    target_rows = [102, 111, 120, 129]
    # 预生成锚定列列表：从 D(4) 开始，步长3，直到 CD(83)
    anchor_cols = list(range(4, 84, 3))  # 4,7,10,...,82
    # CD 是83，但我们取到82为止，因为下一列是85? 实际上CD是83，但步长3，83-4=79，79/3=26.33，所以最后一个小于83的是82（即 CD-1）。可能我们需要直接指定到83。我们改为生成到最大83。
    anchor_cols = list(range(4, 84, 3))  # 4,7,10,...,82，但我们需要包含83，手动调整
    if anchor_cols[-1] != 83:
        anchor_cols.append(83)  # 添加CD列
    # 但这样会多一个，我们实际图片数量可能小于27，所以只取需要的。

    # 4. 收集图片：按目标行分组
    images_by_target = {row: [] for row in target_rows}

    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                src_row = from_cell.row + 1
                src_col = from_cell.col + 1
                # 源区域：C-CG 列，行 101-130
                if 101 <= src_row <= 130 and 3 <= src_col <= 85:
                    # 确定属于哪个目标行（±1 范围）
                    for target_row in target_rows:
                        if target_row - 1 <= src_row <= target_row + 1:
                            images_by_target[target_row].append((src_col, img))
                            break

    # 5. 插入图片
    print("   📝 插入图片...")
    for target_row, imgs in images_by_target.items():
        if not imgs:
            print(f"      ⚠️ 目标行 {target_row} 没有找到图片")
            continue
        # 按列排序
        imgs.sort(key=lambda x: x[0])
        # 获取该行要用的锚定列（取前 len(imgs) 个）
        used_cols = anchor_cols[:len(imgs)]
        for idx, (src_col, img_obj) in enumerate(imgs):
            target_col = used_cols[idx]
            try:
                img_data = img_obj._data()
                img_stream = BytesIO(img_data)
                new_img = XLImage(img_stream)
                col_width_px = get_merged_cell_total_width(ws_target, target_row, target_col)
                row_height_px = get_merged_cell_total_height(ws_target, target_row, target_col)
                new_img.width = col_width_px * 0.99
                new_img.height = row_height_px * 0.99
                ws_target.add_image(new_img, f"{get_column_letter(target_col)}{target_row}")
                print(f"      ✅ 插入图片到 {get_column_letter(target_col)}{target_row}（来自行{src_row}列{get_column_letter(src_col)}）")
            except Exception as e:
                print(f"      ⚠️ 插入图片到 {get_column_letter(target_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Compressed ball diameter 图片更新完成（间隔锚定）")
    print("=" * 60)

def update_compressed_ball_diameter_data(wb, data_folder_path):
    """
    从 IPQC Data 的 Compressed ball diameter Sheet 复制数据到报告
    数据区域：D-AQ 列，行 33-37, 39-43, 45-49, 51-55, 57-61, 63-67, 69-73, 75-79, 81-85
    然后为 E-AQ 列的目标行写入条件公式（IF 原公式=0 则显示空）
    """
    from openpyxl.utils import get_column_letter
    import re

    print("\n" + "=" * 60)
    print("📝 更新 Compressed ball diameter 数据")
    print("=" * 60)

    try:
        ws_target = wb["Compressed ball diameter"]
    except KeyError:
        print("   ⚠️ 未找到 'Compressed ball diameter' Sheet，跳过数据更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Compressed ball diameter" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Compressed ball diameter"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Compressed ball diameter' Sheet，跳过更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # ---------- 1. 复制数据（9 个区域） ----------
    ranges = [
        (33, 37), (39, 43), (45, 49), (51, 55), (57, 61),
        (63, 67), (69, 73), (75, 79), (81, 85)
    ]
    col_start = 4  # D 列
    col_end = 43   # AQ 列

    print("   📝 复制数据（D-AQ 列，9 个区域）...")
    for start_row, end_row in ranges:
        for row in range(start_row, end_row + 1):
            for col in range(col_start, col_end + 1):
                val = ws_source.cell(row=row, column=col).value
                # 数值为 0 时清空
                if val == 0 or val == 0.0:
                    ws_target.cell(row=row, column=col, value=None)
                else:
                    ws_target.cell(row=row, column=col, value=val)
        print(f"      行 {start_row}-{end_row} 更新完成")

    # ---------- 2. 写入条件公式（E-AQ 列，目标行） ----------
    target_rows = [38, 44, 50, 56, 62, 68, 74, 80, 86] + list(range(87, 92))
    print("   📝 写入 E-AQ 列条件公式（目标行），0 值显示为空...")

    # 获取 D 列对应行的公式，并生成替换列字母后的公式
    def adjust_formula_for_column(formula, from_col, to_col):
        """
        将公式中的列字母 from_col 替换为 to_col，同时处理绝对引用 $from_col -> $to_col
        """
        if not formula:
            return formula
        # 先处理绝对引用 $from_col
        formula = formula.replace(f"${from_col}", f"${to_col}")
        # 再处理普通引用
        # 使用正则替换单词边界，避免替换到函数名中的字母（但统计函数中无 D）
        formula = re.sub(rf'\b{from_col}(?=\d)', to_col, formula)
        return formula

    for row in target_rows:
        # 读取 D 列的公式（不含等号）
        d_cell = ws_source.cell(row=row, column=4)
        if d_cell.data_type == 'f':
            base_formula = d_cell.value  # 包含 '='
        else:
            # 如果 D 列不是公式（可能为数值），跳过
            continue

        if not base_formula.startswith('='):
            continue

        # 去掉开头的 '='
        base_formula_no_eq = base_formula[1:]

        # 对于 E 到 AQ 的每一列（列索引 5 到 43）
        for col in range(5, 44):
            from_col_letter = 'D'
            to_col_letter = get_column_letter(col)
            # 调整公式中的列引用
            adjusted_formula = adjust_formula_for_column(base_formula_no_eq, from_col_letter, to_col_letter)
            # 构建条件公式：=IF(调整后公式=0,"",调整后公式)
            new_formula = f"=IF({adjusted_formula}=0,\"\",{adjusted_formula})"
            ws_target.cell(row=row, column=col, value=new_formula)

    wb_ipqc.close()
    print("   ✅ Compressed ball diameter 数据更新完成（含条件公式）")
    print("=" * 60)

def update_peel_strength_images(wb, data_folder_path):
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from io import BytesIO

    print("\n" + "=" * 60)
    print("📝 更新 Peel Strength 图片")
    print("=" * 60)

    try:
        ws_target = wb["Peel Strength"]
    except KeyError:
        print("   ⚠️ 未找到 'Peel Strength' Sheet，跳过图片更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过图片更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Peel Strength" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Peel Strength"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Peel Strength' Sheet，跳过图片更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 1. 清除旧图片
    if ws_target._images:
        images_to_remove = []
        for img in ws_target._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if 5 <= img_row <= 79 and 7 <= img_col <= 30:
                    images_to_remove.append(img)
        for img in images_to_remove:
            ws_target._images.remove(img)
        print(f"      ✅ 已删除 {len(images_to_remove)} 张旧图片")

    # 2. 修复列宽
    print("   📏 修复列宽（S-W, Y-AC 与 S 列一致）...")
    s_col_letter = get_column_letter(19)
    s_width = ws_target.column_dimensions[s_col_letter].width
    if s_width is None:
        s_width = 12
    for col in range(19, 24):
        ws_target.column_dimensions[get_column_letter(col)].width = s_width
    for col in range(25, 30):
        ws_target.column_dimensions[get_column_letter(col)].width = s_width
    print(f"      ✅ S-W 和 Y-AC 列宽已设置为 {s_width} 字符")

    # 3. 定义尺寸计算函数
    def get_merged_cell_total_height(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_height = 0
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    h = ws.row_dimensions[r].height
                    if h is None:
                        h = 15
                    total_height += h
                return total_height * 1.3333
        h = ws.row_dimensions[row].height
        if h is None:
            h = 15
        return h * 1.3333

    def get_merged_cell_total_width(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and \
               col in range(merged_range.min_col, merged_range.max_col + 1):
                total_width = 0
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    w = ws.column_dimensions[get_column_letter(c)].width
                    if w is None:
                        w = 12
                    total_width += w
                return total_width * 8
        w = ws.column_dimensions[get_column_letter(col)].width
        if w is None:
            w = 12
        return w * 8

    # 4. 收集源图片（使用行范围查找）
    target_rows = [7, 14, 26, 33, 45, 52, 64, 71]
    # 对于每个目标行，从源图片中查找范围内的行
    # 我们将所有源图片按列分组，并存储其行和图片对象
    source_images = {}  # key: (col, target_row_index) -> img_obj? 
    # 更简单：对每个目标行，搜索源图片中行在 [target_row-1, target_row+1] 范围内的图片
    # 如果找到多个，则使用行最接近的那一个（或按列处理）
    # 我们按列收集，同一目标行可能有多张图片（不同列）

    # 先收集所有源图片信息
    all_source_imgs = []
    if hasattr(ws_source, '_images') and ws_source._images:
        for img in ws_source._images:
            if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                from_cell = img.anchor._from
                img_row = from_cell.row + 1
                img_col = from_cell.col + 1
                if 5 <= img_row <= 79 and (8 <= img_col <= 17 or 19 <= img_col <= 29):
                    all_source_imgs.append((img_row, img_col, img))

    # 按目标行处理
    for idx, target_row in enumerate(target_rows):
        # 查找源图片中行在 [target_row-1, target_row+1] 范围内的图片
        row_range = (target_row - 1, target_row + 1)
        # 按列分组，对于每一列，选择行最接近 target_row 的图片
        col_best = {}
        for src_row, src_col, img_obj in all_source_imgs:
            if row_range[0] <= src_row <= row_range[1]:
                # 该图片在范围内
                if src_col not in col_best:
                    col_best[src_col] = (src_row, img_obj)
                else:
                    # 如果已有图片，比较哪个行更接近 target_row
                    existing_row = col_best[src_col][0]
                    if abs(src_row - target_row) < abs(existing_row - target_row):
                        col_best[src_col] = (src_row, img_obj)
        # 现在 col_best 包含了该目标行所有列的最佳图片
        # 对每个列插入图片
        for src_col, (src_row, img_obj) in col_best.items():
            try:
                img_data = img_obj._data()
                img_stream = BytesIO(img_data)
                new_img = XLImage(img_stream)
                col_width_px = get_merged_cell_total_width(ws_target, target_row, src_col)
                row_height_px = get_merged_cell_total_height(ws_target, target_row, src_col)
                new_img.width = col_width_px * 0.99
                new_img.height = row_height_px * 0.99
                ws_target.add_image(new_img, f"{get_column_letter(src_col)}{target_row}")
                print(f"      ✅ 插入图片到 {get_column_letter(src_col)}{target_row}（来自行{src_row}）")
            except Exception as e:
                print(f"      ⚠️ 插入图片到 {get_column_letter(src_col)}{target_row} 失败：{e}")

    wb_ipqc.close()
    print("   ✅ Peel Strength 图片更新完成（使用行范围查找）")
    print("=" * 60)

def update_peel_strength_data(wb, data_folder_path):
    """
    从 IPQC Data 的 Peel Strength Sheet 复制 E-F 列数据到报告，
    并在 E83:H92 写入带条件判断的公式（隐藏零值）。
    数据区域：行 6-15, 25-34, 44-53, 63-72 (E-F 列)
    公式区域：E83:E92 引用 F6:F15；F83:F92 引用 F25:F34；
              G83:G92 引用 F44:F53；H83:H92 引用 F63:F72
    """
    print("\n" + "=" * 60)
    print("📝 更新 Peel Strength 数据")
    print("=" * 60)

    try:
        ws_target = wb["Peel Strength"]
    except KeyError:
        print("   ⚠️ 未找到 'Peel Strength' Sheet，跳过数据更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Peel Strength" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Peel Strength"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Peel Strength' Sheet，跳过更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    # 数据区域：行范围，列 E-F (5-6)
    ranges = [
        (6, 15),
        (25, 34),
        (44, 53),
        (63, 72)
    ]

    print("   📝 复制数据（E-F 列）...")
    for start_row, end_row in ranges:
        for row in range(start_row, end_row + 1):
            for col in range(5, 7):  # E=5, F=6
                val = ws_source.cell(row=row, column=col).value
                ws_target.cell(row=row, column=col, value=val)
        print(f"      行 {start_row}-{end_row} 更新完成")

    wb_ipqc.close()

    # ============================================================
    # 【新增】写入 E83:H92 的公式（隐藏零值）
    # ============================================================
    print("   📝 写入 E83:H92 公式（隐藏零值）...")

    # E83:E92 = IF(F6=0,"",F6) 到 IF(F15=0,"",F15)
    for i, row in enumerate(range(83, 93)):
        src_row = 6 + i
        ws_target.cell(row=row, column=5, value=f"=IF(F{src_row}=0,\"\",F{src_row})")

    # F83:F92 = IF(F25=0,"",F25) 到 IF(F34=0,"",F34)
    for i, row in enumerate(range(83, 93)):
        src_row = 25 + i
        ws_target.cell(row=row, column=6, value=f"=IF(F{src_row}=0,\"\",F{src_row})")

    # G83:G92 = IF(F44=0,"",F44) 到 IF(F53=0,"",F53)
    for i, row in enumerate(range(83, 93)):
        src_row = 44 + i
        ws_target.cell(row=row, column=7, value=f"=IF(F{src_row}=0,\"\",F{src_row})")

    # H83:H92 = IF(F63=0,"",F63) 到 IF(F72=0,"",F72)
    for i, row in enumerate(range(83, 93)):
        src_row = 63 + i
        ws_target.cell(row=row, column=8, value=f"=IF(F{src_row}=0,\"\",F{src_row})")
    # ============================================================
    # 【新增】写入 E105:H114 的公式（隐藏零值，引用 E 列数据）
    # ============================================================
    print("   📝 写入 E105:H114 公式（隐藏零值）...")

    # E105:E114 = IF(E6=0,"",E6) 到 IF(E15=0,"",E15)
    for i, row in enumerate(range(105, 115)):
        src_row = 6 + i
        ws_target.cell(row=row, column=5, value=f"=IF(E{src_row}=0,\"\",E{src_row})")

    # F105:F114 = IF(E25=0,"",E25) 到 IF(E34=0,"",E34)
    for i, row in enumerate(range(105, 115)):
        src_row = 25 + i
        ws_target.cell(row=row, column=6, value=f"=IF(E{src_row}=0,\"\",E{src_row})")

    # G105:G114 = IF(E44=0,"",E44) 到 IF(E53=0,"",E53)
    for i, row in enumerate(range(105, 115)):
        src_row = 44 + i
        ws_target.cell(row=row, column=7, value=f"=IF(E{src_row}=0,\"\",E{src_row})")

    # H105:H114 = IF(E63=0,"",E63) 到 IF(E72=0,"",E72)
    for i, row in enumerate(range(105, 115)):
        src_row = 63 + i
        ws_target.cell(row=row, column=8, value=f"=IF(E{src_row}=0,\"\",E{src_row})")


    print("   ✅ E83:H92 公式写入完成")
    print("   ✅ E105:H114 公式写入完成")
    print("   ✅ Peel Strength 数据更新完成")
    print("=" * 60)

def update_placement_accuracy_report(wb, data_folder_path):
    """
    从 IPQC Data 的 Placement accuracy Sheet 复制指定区域到报告中，
    完全覆盖目标区域（包括空值），模拟手动复制-粘贴操作。
    区域：E7:G16, E27:G36, E47:G56, E67:G76
    """
    print("\n" + "=" * 60)
    print("📝 更新 Placement accuracy Sheet（完全覆盖）")
    print("=" * 60)

    try:
        ws_target = wb["Placement accuracy"]
    except KeyError:
        print("   ⚠️ 未找到 'Placement accuracy' Sheet，跳过更新")
        return

    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    ipqc_file = get_first_file_in_folder(ipqc_folder, extensions=['.xlsx', '.xlsm', '.xls'])
    if not ipqc_file:
        print("   ⚠️ IPQC Data 文件夹中没有 Excel 文件，跳过更新")
        return

    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Placement accuracy" in wb_ipqc.sheetnames:
            ws_source = wb_ipqc["Placement accuracy"]
        else:
            print("   ⚠️ IPQC Data 文件中没有 'Placement accuracy' Sheet，跳过更新")
            wb_ipqc.close()
            return
    except Exception as e:
        print(f"   ❌ 读取 IPQC Data 文件失败：{e}")
        return

    ranges = [
        (7, 16),
        (27, 36),
        (47, 56),
        (67, 76)
    ]

    print("   📝 复制数据（完全覆盖，包括空值）...")
    for start_row, end_row in ranges:
        for row in range(start_row, end_row + 1):
            for col in range(5, 8):  # E=5, F=6, G=7
                val = ws_source.cell(row=row, column=col).value
                ws_target.cell(row=row, column=col, value=val)
        print(f"      行 {start_row}-{end_row} 更新完成")

    wb_ipqc.close()
    print("   ✅ Placement accuracy 更新完成（完全覆盖）")
    print("=" * 60)

def update_acf_summary_report(template_path, data_folder_path, output_folder_path=None):
    print("\n" + "=" * 60)
    print("📝 更新 ACF Summary Sheet")
    print("=" * 60)

    today = datetime.now()
    today_str = today.strftime("%Y%m%d")
    today_slash = today.strftime("%Y/%m/%d")

    folder_name = os.path.basename(data_folder_path)
    parsed = parse_folder_name(folder_name)
    if not parsed:
        return
    专案 = parsed["专案"]
    阶段 = parsed["阶段"]
    Config = parsed["Config"]
    机台号 = parsed["机台号"]
    报告类型 = parsed["报告类型"]

    print(f"📁 数据文件夹：{folder_name}")
    print(f"   📌 专案：{专案}")
    print(f"   📌 阶段：{阶段}")
    print(f"   📌 Config：{Config}")
    print(f"   📌 机台号：{机台号}")
    print(f"   📌 报告类型：{报告类型}")

    report_name = f"{专案}_ABU_{Config}_{报告类型}_ACF_{机台号}_{today_str}_Rev.0"
    print(f"📄 新报告名称：{report_name}")

    if not os.path.exists(template_path):
        print(f"❌ 错误：找不到模板文件 {template_path}")
        return
    if output_folder_path is None:
        output_folder_path = data_folder_path
    output_path = os.path.join(output_folder_path, f"{report_name}.xlsx")
    shutil.copy2(template_path, output_path)
    print(f"✅ 已复制模板到：{output_path}")

    # ============================================================
    # 【关键】先以 data_only=True 读取 I 列的数值（公式结果）
    # ============================================================
    print("📝 正在读取 I 列数值（公式结果）...")
    temp_wb = openpyxl.load_workbook(output_path, data_only=True)
    temp_ws = temp_wb.active
    i_values = {}
    for row in range(27, 193):
        i_values[row] = temp_ws.cell(row=row, column=9).value
    for row in range(196, 218):
        i_values[row] = temp_ws.cell(row=row, column=9).value
    i_values[26] = temp_ws.cell(row=26, column=9).value
    i_values[195] = temp_ws.cell(row=195, column=9).value
    temp_wb.close()
    print("   ✅ I 列数值已缓存")

    # ============================================================
    # 现在以默认方式加载工作簿进行修改
    # ============================================================
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # ============================================================
    # 【第一步】智能确定当前数据列范围
    # ============================================================
    print("📝 正在检测当前数据列范围...")
    merged = get_merged_cell_range(ws, 26, 11)  # K=11
    if merged:
        start_col = merged[2]
        end_col = merged[3]
        print(f"   ✅ 检测到合并单元格，列范围：{get_column_letter(start_col)} ~ {get_column_letter(end_col)}")
    else:
        start_col = 11
        end_col = start_col
        while True:
            next_col = end_col + 1
            val = ws.cell(row=26, column=next_col).value
            if val is None:
                break
            end_col = next_col
        print(f"   ✅ 检测到连续非空单元格，列范围：{get_column_letter(start_col)} ~ {get_column_letter(end_col)}")

    col_count = end_col - start_col + 1
    print(f"   📊 当前数据列数：{col_count}")

    # ============================================================
    # 【第二步】历史数据归档：整体右移1列
    # ============================================================
    print("📝 正在执行历史数据归档（整体右移1列）...")
    for row in range(27, 193):
        for col in range(end_col, start_col - 1, -1):
            val = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col + 1, value=val)
            if col > start_col:
                safe_set_cell_value(ws, row, col, None)

    for row in range(196, 218):
        for col in range(end_col, start_col - 1, -1):
            val = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col + 1, value=val)
            if col > start_col:
                safe_set_cell_value(ws, row, col, None)

    for row in [26, 195]:
        for col in range(end_col, start_col - 1, -1):
            val = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col + 1, value=val)
            if col > start_col:
                safe_set_cell_value(ws, row, col, None)

    print("   ✅ 数据右移完成")

    # ============================================================
    # 【第三步】将缓存的 I 列数值写入 K 列（只写数值，不带公式）
    # ============================================================
    print("📝 正在将 I 列数值（不含公式）写入 K 列...")
    for row, val in i_values.items():
        # 跳过 K26 和 K195，它们应为 Previous Config
        if row == 26 or row == 195:
            continue
        # 检查是否为 #DIV/0! 错误
        if isinstance(val, str) and val.startswith("#DIV/0!"):
            val = "NA"
        ws.cell(row=row, column=start_col, value=val)

    # 设置 K26 和 K195 为 Previous Config
    ws.cell(row=26, column=start_col, value="Previous Config")
    ws.cell(row=195, column=start_col, value="Previous Config")
    print("   ✅ I 列数值已粘贴到 K 列（仅数值，不含公式），K26/K195 已设为 Previous Config")

    # ============================================================
    # 【第四步】合并 K26 和 K195 到当前数据列范围
    # ============================================================
    print("📝 正在合并 K26 和 K195 到数据列范围...")
    to_unmerge = []
    for merged in ws.merged_cells.ranges:
        if (merged.min_row == 26 and merged.min_col >= start_col and merged.max_col <= end_col) or \
           (merged.min_row == 195 and merged.min_col >= start_col and merged.max_col <= end_col):
            to_unmerge.append(merged)
    for merged in to_unmerge:
        ws.unmerge_cells(str(merged))

    ws.merge_cells(start_row=26, start_column=start_col, end_row=26, end_column=end_col)
    ws.cell(row=26, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=195, start_column=start_col, end_row=195, end_column=end_col)
    ws.cell(row=195, column=start_col).alignment = Alignment(horizontal='center', vertical='center')
    print("   ✅ 合并完成")

    # ============================================================
    # 【第五步】填充基本信息
    # ============================================================
    print("📝 填充基本信息...")
    safe_set_cell_value(ws, 5, 5, 专案)        # E5
    safe_set_cell_value(ws, 9, 8, 报告类型)    # H9: Event (MBO/PBO)
    safe_set_cell_value(ws, 4, 8, today_slash) # H4

    # 版本号提取
    ers_file = get_first_file_in_folder(os.path.join(data_folder_path, "ERS"))
    if ers_file:
        ers_rev = extract_revision_from_file(ers_file, r'Table\s*:\s*Process\s*Control\s*Rev\s*([0-9]+\.[0-9]+)')
        if ers_rev:
            safe_set_cell_value(ws, 5, 16, ers_rev)
            print(f"   ✅ P5 = {ers_rev}")
        else:
            print("   ⚠️ 未找到 Process Control Rev")

        try:
            with open(ers_file, 'r', encoding='utf-8', errors='ignore') as f:
                ers_content = f.read()
            head = ers_content[:2000]
            match = re.search(r'Rev\s*([0-9]+\.[0-9]+)', head, re.IGNORECASE)
            if match:
                safe_set_cell_value(ws, 5, 17, match.group(1))
                print(f"   ✅ Q5 = {match.group(1)}")
        except:
            pass

    vsr_file = get_first_file_in_folder(os.path.join(data_folder_path, "VSR"))
    if vsr_file:
        vsr_filename = os.path.basename(vsr_file)
        match = re.search(r'Rev\s*([0-9]+)', vsr_filename)
        if match:
            safe_set_cell_value(ws, 5, 18, match.group(1))
            print(f"   ✅ R5 = {match.group(1)}")
        else:
            try:
                with open(vsr_file, 'r', encoding='utf-8', errors='ignore') as f:
                    vsr_content = f.read()
                match = re.search(r'Rev\s*([0-9]+)', vsr_content[:2000], re.IGNORECASE)
                if match:
                    safe_set_cell_value(ws, 5, 18, match.group(1))
                    print(f"   ✅ R5 = {match.group(1)}")
            except:
                pass

    mco_file = get_first_file_in_folder(os.path.join(data_folder_path, "MCO"))
    if mco_file:
        mco_filename = os.path.basename(mco_file)
        match = re.search(r'[0-9]+-[0-9]+-([0-9]+)', mco_filename)
        if match:
            safe_set_cell_value(ws, 5, 19, match.group(1))
        else:
            parts = mco_filename.split('-')
            if len(parts) >= 3:
                safe_set_cell_value(ws, 5, 19, parts[2].split('.')[0])
        print(f"   ✅ S5 = {ws.cell(row=5, column=19).value}")

    print("✅ 基本信息填充完成")
    
    # 更新 Placement accuracy（完全覆盖）
    update_placement_accuracy_report(wb, data_folder_path)
    # 更新 Peel Strength
    update_peel_strength_data(wb, data_folder_path)
    update_peel_strength_images(wb, data_folder_path)
    # 更新 Compressed ball diameter
    update_compressed_ball_diameter_data(wb, data_folder_path)
    update_compressed_ball_diameter_images(wb, data_folder_path)
    # 更新第二个区块
    update_compressed_ball_diameter_block2(wb, data_folder_path)
    # 更新 Lamination placement
    update_lamination_placement_report(wb, data_folder_path)
    update_lamination_placement_images(wb, data_folder_path)
    # 更新 ACF squeeze out
    update_acf_squeeze_out_report(wb, data_folder_path)
    # 更新 Break Mode
    update_break_mode_report(wb, data_folder_path)
    # 更新 FAI
    update_fai_report(wb, data_folder_path)
    # 更新 Module integrity
    update_module_integrity_report(wb, data_folder_path, 报告类型)
    # 保存 openpyxl 的所有修改（数据部分）
    wb.save(output_path)


    print("=" * 60)
    print(f"✅ ACF Summary 更新完成！文件：{output_path}")
    print("=" * 60)
    return output_path

# ============================================================
# 入口
# ============================================================

if __name__ == "__main__":
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
    except:
        print("⚠️ tkinter 不可用，请手动输入路径。")
        data_folder = input("请输入数据文件夹路径：").strip()
        template_file = input("请输入模板文件路径：").strip()
    else:
        data_folder = filedialog.askdirectory(
            title="请选择数据文件夹（命名格式：专案 阶段 Config 机台号 M/PBO）"
        )
        if not data_folder:
            print("❌ 未选择数据文件夹，程序退出")
            sys.exit(0)

        template_file = filedialog.askopenfilename(
            title="请选择报告模板文件（.xlsx格式）",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not template_file:
            print("❌ 未选择模板文件，程序退出")
            sys.exit(0)

    folder_name = os.path.basename(data_folder)
    if not parse_folder_name(folder_name):
        if messagebox:
            messagebox.showerror(
                "格式错误",
                f"数据文件夹名称格式不正确！\n应为：专案 阶段 Config 机台号 M/PBO\n当前：{folder_name}"
            )
        else:
            print(f"❌ 文件夹名称格式不正确：{folder_name}")
        sys.exit(1)

    update_acf_summary_report(template_file, data_folder)